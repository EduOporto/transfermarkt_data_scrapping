import requests
from bs4 import BeautifulSoup
import bs4
import pandas as pd 
import itertools
import math
import os
from openpyxl import load_workbook
from os import path
import tqdm
import time
import subprocess

def player_search_results(page,user_search, user_search_for_link):
#Soup
    query_link = 'https://www.transfermarkt.com/schnellsuche/ergebnis/schnellsuche?Spieler_page={}&ajax=yw0&query={}&x=0&y=0'.format(page,user_search_for_link)
    search_page = requests.get(query_link, headers={'User-Agent': 'Mozilla/5.0'})
    search_soup = BeautifulSoup(search_page.content, 'lxml')
#Scrape the data to build the lists
    players_names = []
    players_pos_age = []
    players_team = []
    players_nat = []
    players_id = []
    hints = []
    for table in search_soup.find_all('div', {'class': 'box'}):
        for e in table.find_all('div', {'class': 'table-header'}):
            if 'players' in e.text:
                for e in e.text:
                    if e != ' ':
                        hints.append(e)
                for e in table.find_all('td', {'class': 'hauptlink'}):
                    if e.attrs == {'class': ['hauptlink']}:
                        for e1 in e:
                            if isinstance(e1, bs4.element.Tag):
                                players_names.append(e1['title'])
                                players_id.append(e1['id'])
                for e in table.find_all('table', {'class':'inline-table'}):
                    for e1 in e.find_all('tr'):
                        players_team.append(e1.text)
                for e in table.find_all('td', {'class': 'zentriert'}):
                    if e.text != '':
                        players_pos_age.append(e.text)
                    for e1 in e.find_all('img'):
                        players_nat.append(e1['title'])
    players_nat.append('\xa0')
    nations = []
    nations_together = []
    for e in players_nat[1:]:
        if e != '\xa0' and e != 'Retired' and e != 'Unknown' and e != 'Without Club':
            nations.append(e)
        if e == '\xa0' or e == 'Retired' or e == 'Unknown' or e == 'Without Club':
            nations_together.append(nations)
            nations = []
                #If there's some problem with the nations, answer here -> print(nations_together)
    players_team = players_team[1::2]
    players_pos = players_pos_age[::2]
    players_age = players_pos_age[1::2]
    index_for_df = list(range(1,len(players_names)+1))
#Merge the lists on a dictionary
    dict_for_df = {'Result Nº':index_for_df,'Player':players_names,'Pos':players_pos,'Team':players_team,'Age':players_age,'Nat':nations_together,'ID':players_id}
    lists_dataframe = pd.DataFrame(dict_for_df) 
    lists_dataframe.set_index('Result Nº', inplace=True)       
#Show hints and options to the user
    hints = int(''.join(hints[hints.index('-')+1:hints.index('H')]))
    pages = math.ceil(hints/10)
    print('\n')
    print("{} hints found for the search '{}'. Showing page {} of {}".format(hints, user_search, page, pages))
    print('\n')
    print(lists_dataframe)
    print('\n')
    checker = str(input('Did I find the player you are looking for(Y/N): '))
    if checker == 'Y':
        result_num = int(input('Which result number is the player you are looking for?: '))
        player_name_clean = lists_dataframe['Player'][result_num]
        player_name_link = player_name_clean.lower().replace(' ', '+')
        player_id = lists_dataframe['ID'][result_num]
        player_season(player_name_link, player_name_clean, player_id)
    if checker == 'N':
        if page != pages:
            more_results = str(input('Would you like to see more results(Y/N)?: '))
            if more_results == 'Y':
                page = page+1
                player_search_results(page,user_search, user_search_for_link)
            if more_results == 'N':
                back_or_new = str(input('Would you like to go back to the first page or make a new search(First page/New search)?: '))
                if back_or_new == 'First page':
                    page = 1
                    player_search_results(page,user_search, user_search_for_link)
                if back_or_new == 'New search':
                    new_search()
        if page == pages:
            print('We already check all the hints')
            back_or_new = str(input('Would you like to go back to the first page or make a new search(First page/New search)?: '))
            if back_or_new == 'First page':
                page = 1
                player_search_results(page,user_search, user_search_for_link)
            if back_or_new == 'New search':
                new_search()

def player_season(player_name_link, player_name_clean, player_id):
#Soup
    link = 'https://www.transfermarkt.com/{}/leistungsdatendetails/spieler/{}/saison//verein/0/liga/0/wettbewerb//pos/0/trainer_id/0/plus/1'.format(player_name_link, player_id)
    search_page = requests.get(link, headers={'User-Agent': 'Mozilla/5.0'})
    search_soup = BeautifulSoup(search_page.content, 'lxml')        
#Get the lists for showing all the seasons the player has been active
    values_list = []
    for e in search_soup.find_all('select', {'data-placeholder':'Filter by season'}):
        seasons_list = e.text.split('\n')
        seasons_list = seasons_list[1:-1]
        for e1 in e:
            if isinstance(e1, bs4.element.Tag):
                values_list.append(e1['value'])
    seasons_index = list(range(1,len(values_list)+1))
#Merge the lists in a dictionary
    dict_for_df = {'Index':seasons_index, 'Seasons':seasons_list, 'Val.':values_list}
    lists_dataframe = pd.DataFrame(dict_for_df) 
    lists_dataframe.set_index('Index', inplace=True)
#Show hints and options to the user
    print('\n')
    print('Here is a frame with all the seasons {} have played along his career:'.format(player_name_clean))
    print('\n')
    print(lists_dataframe)
    print('\n')
    choice = str(input('Would you like to see any particular season or download the whole career on Excel(Season/Career)?: '))
    print('\n')
    if choice == 'Season':
        auto_saver = 'No'
        season_choice = int(input("Which season's data would you like to know (pick the index number)? "))
        season = lists_dataframe['Val.'][season_choice]
        season_clean = lists_dataframe['Seasons'][season_choice]
        if season_choice > 1:
            transfermarkt_scrapper_season(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe)
        if season_choice == 1:
            transfermarkt_scrapper_career(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe)
    if choice == 'Career':
        auto_saver = 'Yes'
        print("Saving {}'s data...".format(player_name_clean))
        for i, season, season_clean in zip(tqdm.tqdm(range(len(lists_dataframe['Seasons']))), lists_dataframe['Val.'], lists_dataframe['Seasons']):
            if season == '':
                transfermarkt_scrapper_career(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe)
            else:
                transfermarkt_scrapper_season(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe)
            time.sleep(0.1)
        print('\n')
        print('The data of {} has been downloaded in {}/Excel workbooks'.format(player_name_clean,os.getcwd()))
        open_directory = str(input('Would you like to open the file directory(Y/N)?: '))
        if open_directory == 'Y':
            file_out_path = '{}/Excel workbooks/{}.xlsx'.format(os.getcwd(), player_name_clean)
            subprocess.call(["open", "-R", file_out_path])
            pass
        if open_directory == 'N':
            pass
        user_choice = str(input('Would you like to do a new search or exit the program (New search/Exit)?: '))
        if user_choice == 'New search':
            new_search()
        if user_choice == 'Exit':
            print('See you next time, thank you!')
            exit

def show_seasons_again(lists_dataframe, player_name_link, player_id, season, player_name_clean, season_clean, auto_saver):
    print('\n')
    print(lists_dataframe)
    print('\n')
    season_choice = int(input("Which season's data would you like to know (pick the index number)? "))
    season = lists_dataframe['Val.'][season_choice]
    season_clean = lists_dataframe['Seasons'][season_choice]
    if season_choice > 1:
        transfermarkt_scrapper_season(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe)
    if season_choice == 1:
        transfermarkt_scrapper_career(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe)

def transfermarkt_scrapper_season(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe):
#Soup
    url = 'https://www.transfermarkt.com/{}/leistungsdatendetails/spieler/{}/saison/{}/verein/0/liga/0/wettbewerb//pos/0/trainer_id/0/plus/1'.format(player_name_link, player_id, season)
    page = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    soup = BeautifulSoup(page.content, 'lxml')
#Check if data of the season
    if len(soup.find_all('span',{'class':'empty'})) > 0:
        if auto_saver == 'No':
            print('There is no data of {} for the season you have chosen'.format(player_name_clean))
            user_excel = str(input('Would you like to save the dataframe as an Excel document anyway(Y/N)?: '))
            if user_excel == 'Y':
                general_df = pd.DataFrame({'Player data':['No information']})
                excel_saver(general_df, player_name_clean, season_clean)
            if user_excel == 'N':
                pass
            more_seasons = str(input('Would you like to see more seasons of {} (Y/N)?: '.format(player_name_clean)))
            if more_seasons == 'Y':
                show_seasons_again(lists_dataframe, player_name_link, player_id, season, player_name_clean, season_clean, auto_saver)
            if more_seasons == 'N':
                user_choice = str(input('Would you like to do a new search or exit the program (New search/Exit)?: '))
                if user_choice == 'New search':
                    new_search()
                if user_choice == 'Exit':
                    print('See you next time, thank you!')
                    exit
        if auto_saver == 'Yes':
            general_df = pd.DataFrame({'Player data':['No information']})
            excel_saver(general_df, player_name_clean, season_clean)
    if len(soup.find_all('span',{'class':'empty'})) == 0:
#Get all the table titles
        list_of_dictionaries = []
        for table_title in soup.find_all('div', {'class': 'table-header img-vat'}):
            list_of_dictionaries.append({table_title.text.strip():{}})
#Get the table headers
            #Get the table headers (1/3)
        for table_header1 in soup.find_all('thead')[1:2]:
            table_header1_list = table_header1.text.strip()
            columns_as_list = table_header1_list.split('\n')
            #Get the table headers (2/3)
        for table_header2 in soup.findAll('th',{'class':'zentriert'})[17:25]:
            table_header2_2 = table_header2.find('span')
            columns_as_list.append(table_header2_2['title'])    
            #Get the table headers (3/3)
        for table_header3 in soup.findAll('th',{'class':'rechts'})[2:3]:
            table_header3_3 = table_header3.find('span')
            columns_as_list.append(table_header3_3['title'])
#Merge the table names with the headers
        competition_list = []
        for e in list_of_dictionaries:
            for comp, table in e.items():
                competition_list.append(comp)
                for h in columns_as_list:
                    table[h] = 0
#Get dataframe elements
        table_data_list = []
        for table_data in soup.find_all('tbody')[2:]:
            table_data_list.append(table_data.text.split('\n'))
        #Get the opponents
        table_data_dict = {}
        for table_data in soup.find_all('tbody')[2:]:
            for opp in table_data.find_all('td', {'class':'no-border-links hauptlink'}):
                table_data_dict[opp.text] = opp.text + ' (opp)'
        #Clean empty spaces at the start and end of every dataframe
        table_data_list_clean = []
        for dataframe in table_data_list:
            del dataframe[:4]
            del dataframe[-2:]
            table_data_list_clean.append(dataframe)
        #Split every dataset in rows and wrap it in a list
        dataframes_splited = []
        dataframes_num = len(table_data_list_clean)
        for df in range(dataframes_num):
            split_list = [i for i, s in enumerate(table_data_list_clean[df]) if '\t' in s]
            list_splitted = [table_data_list_clean[df][i:j] for i, j in zip([df] + split_list, split_list + [None])]
            del list_splitted[0]
            dataframes_splited.append(list_splitted)
        #Delete leftovers
        dataframes_splited_cleaned_1 = []
        for e in dataframes_splited:
            rows_cleaned = []
            for e1 in e:
                no_tabs = ''.join(e1[0].split())
                del e1[0]
                del e1[1]
                del e1[2]
                e1.insert(0,no_tabs)
                rows_cleaned.append(e1)
            dataframes_splited_cleaned_1.append(rows_cleaned)
        dataframes_splited_cleaned_final = []
        for e in dataframes_splited_cleaned_1:
            rows_4_spaces = e[:-1]
            rows_cleaned = []
            for e1 in rows_4_spaces:
                rows_without_4_spaces = e1[:-4]
                rows_cleaned.append(rows_without_4_spaces)
            rows_cleaned.append(e[-1])
            dataframes_splited_cleaned_final.append(rows_cleaned)
        #Add spaces when the player did not play a match
        for e in dataframes_splited_cleaned_final:
            for e1 in e:
                if len(e1) != 15:
                    to_add = list(itertools.repeat('', 15-len(e1)))
                    for s in to_add:
                        e1.append(s)
        #Add (opp) to the opponents
        for comp in dataframes_splited_cleaned_final:
            for match in comp:
                for n, e in enumerate(match):
                    for key, value in table_data_dict.items():
                        if e == key:
                            match[n] = value
#Join the dataframes with the data
        datasets_dict = {}
        competiton_and_data_dict = dict(zip(competition_list,dataframes_splited_cleaned_final))
        for competition, data in competiton_and_data_dict.items():
            for comp in competition_list:
                if competition == comp:
                    columns_together = []
                    for e in range(len(columns_as_list)):
                        column_data = []
                        for row in data:
                            column_data.append(row[e])
                        columns_together.append(column_data)
                    dataframe_order_dict = dict(zip(columns_as_list,columns_together))
                    datasets_dict[competition] = dataframe_order_dict
#Convert the data into dataframes and concatenate them all together
        list_final_dfs = []
        for competition, dataset in datasets_dict.items():
            comp_dataframe = pd.DataFrame.from_dict(dataset)
            dataframe_len = len(comp_dataframe.index)
            comp_dataframe.insert(0, 'Competition', list(itertools.repeat(competition, dataframe_len)), True)
            list_final_dfs.append(comp_dataframe)
        general_df = pd.concat(list_final_dfs, ignore_index=True)
#Show the results and options for the user
        results_and_options(auto_saver, general_df, player_name_clean, season_clean, lists_dataframe, player_name_link, player_id, season)

def transfermarkt_scrapper_career(player_name_link, player_id, season, player_name_clean, season_clean, auto_saver, lists_dataframe):
#Soup
    url = 'https://www.transfermarkt.com/{}/leistungsdatendetails/spieler/{}/saison/{}/verein/0/liga/0/wettbewerb//pos/0/trainer_id/0/plus/1'.format(player_name_link, player_id, season)
    page = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    soup = BeautifulSoup(page.content, 'lxml')
#Get all the table headers
    headers_1 = []
    headers_2 = []
    for table_header1 in soup.find_all('th'):
        if table_header1.text != 'reihenfolge ASC, saison_id DESC, verein_id ASC':
            headers_1.append(table_header1.text)
    for table_header2 in soup.findAll('th',{'class':'zentriert'}):
        table_header2_2 = table_header2.find('span')
        if table_header2_2 != None:
            headers_2.append(table_header2_2['title'])
    for table_header3 in soup.findAll('th',{'class':'rechts'}):
        table_header3_3 = table_header3.find('span')
        headers_2.append(table_header3_3['title'])
#Merge all the table headers
    for h2 in headers_2:
        for n, h1 in enumerate(headers_1):
            if h1 == '\xa0':
                headers_1[n] = h2
    headers = headers_1[:6] + headers_2[1:]
#Get dataframe elements
    table_data_zentriert = []
    table_data_hauptlink = []
    table_data_rechts = []
    teams = []
    for table_data in soup.find_all('tbody')[1:]:
        for table_row in table_data.find_all('td', {'class':'zentriert'}):
            table_data_zentriert.append(table_row.text)
        for table_row in table_data.find_all('td', {'class':'hauptlink'}):
            table_data_hauptlink.append(table_row.text)
        for table_row in table_data.find_all('td', {'class':'rechts'}):
            table_data_rechts.append(table_row.text)
        for table_row in table_data.find_all('a', {'class':'vereinprofil_tooltip'}):
            for e in table_row.find_all('img'):
                teams.append(e.get('alt'))
#Clean dataframe elements
    table_data_zentriert = [table_data_zentriert[i * 14:(i + 1) * 14] for i in range((len(table_data_zentriert) + 14 -1) // 14)]
    table_data_hauptlink = [x for x in table_data_hauptlink if x != '']
    table_data_rechts = [table_data_rechts[i * 2:(i + 1) * 2] for i in range((len(table_data_rechts) + 2 -1) // 2)]
#Merge cleaned dataframe headers and elements in a dictionary
    rows_final = []
    for row, team in zip(table_data_zentriert, teams):
        row.insert(2, team)
    for row, comp in zip(table_data_zentriert, table_data_hauptlink):
        row[1] = comp
    for row1, row2 in zip(table_data_zentriert, table_data_rechts):
        rows_final.append(row1 + row2)
    columns_together = []
    for e in range(len(headers)):
        column_data = []
        for row in rows_final:
            column_data.append(row[e])
        columns_together.append(column_data)
    dataframe_order_dict = dict(zip(headers, columns_together))
    career_dataframe = pd.DataFrame(dataframe_order_dict)
#Show the results and options for the user
    results_and_options(auto_saver, career_dataframe, player_name_clean, season_clean, lists_dataframe, player_name_link, player_id, season)

def results_and_options(auto_saver, general_df, player_name_clean, season_clean, lists_dataframe, player_name_link, player_id, season):
    if auto_saver == 'No':
        print('\n')
        print(general_df)
        print('\n')
        user_excel = str(input('Would you like to save the dataframe as an Excel document(Y/N)?: '))
        if user_excel == 'Y':
            print("Saving {}'s data...".format(player_name_clean))
            for i in tqdm.tqdm(range(1)):
                excel_saver(general_df, player_name_clean, season_clean)
                time.sleep(0.1)
            print('\n')
            print('The data of {} has been downloaded in {}/Excel workbooks'.format(player_name_clean,os.getcwd()))
            open_directory = str(input('Would you like to open the file directory(Y/N)?: '))
            if open_directory == 'Y':
                file_out_path = '{}/Excel workbooks/{}.xlsx'.format(os.getcwd(), player_name_clean)
                subprocess.call(["open", "-R", file_out_path])
                pass
            if open_directory == 'N':
                pass
        if user_excel == 'N':
            pass
        more_seasons = str(input('Would you like to see more seasons of {} (Y/N)?: '.format(player_name_clean)))
        if more_seasons == 'Y':
            show_seasons_again(lists_dataframe, player_name_link, player_id, season, player_name_clean, season_clean, auto_saver)
        if more_seasons == 'N':
            user_choice = str(input('Would you like to do a new search or exit the program (New search/Exit)?: '))
            if user_choice == 'New search':
                new_search()
            if user_choice == 'Exit':
                print('See you next time, thank you!')
                exit
    if auto_saver == 'Yes':
        excel_saver(general_df, player_name_clean, season_clean)

def excel_saver(dataframe, player_name_clean, season_clean):
    out_path = '{}/Excel workbooks'.format(os.getcwd())
    season_clean = season_clean.replace('/', '-')
    if path.exists('{}/{}.xlsx'.format(out_path, player_name_clean)):
        wb = load_workbook('{}/{}.xlsx'.format(out_path, player_name_clean))
        if season_clean in wb.sheetnames:
            print('The season you trying to save has been already saved in the document')
        else:
            writer = pd.ExcelWriter('{}/{}.xlsx'.format(out_path, player_name_clean), engine= 'openpyxl')
            writer.book = wb
            dataframe.to_excel(writer, sheet_name='{}'.format(season_clean))
            writer.save()
            writer.close()
    else:
        dataframe.to_excel(r'{}/{}.xlsx'.format(out_path, player_name_clean), sheet_name='{}'.format(season_clean))

def new_search():
    print('\n')
    user_search = input("Which player's data would you like to know?: ")
    user_search_for_link = user_search.lower().replace(' ', '+')
    page = 1
    player_search_results(page,user_search, user_search_for_link)

##########

print('\n')
print("Welcome to the 'Transfermarkt Dataset Scrapper'.\nWith this program you will be able to download the historical records of all the football players\nTransfermarkt.com has registered, either from a particular season or on his whole career, in Excel format")
print('\n')
print("To do so, you just have to perfrom a search, select the player you want and choose between the two options.\nAll the players will be saved in the 'Excel workbooks' folder. Let's start!")
new_search()